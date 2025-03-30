from io import BytesIO
from os import getenv

from dotenv import load_dotenv
from flask import Flask, jsonify, request, Response
from openpyxl.reader.excel import load_workbook
from openpyxl.styles.borders import Border

load_dotenv()

app = Flask(__name__)
allowed_addresses = {"127.0.0.1", "localhost", "::1"}


@app.before_request
def limit_remote_addr() -> tuple[Response, int] | None:
    if request.remote_addr not in allowed_addresses:
        return jsonify({"error": "Access denied."}), 403


@app.route("/udec-bot/api/parse-xlsx-borders", methods=["POST"])
def parse_xlsx_borders() -> tuple[Response, int]:
    if "file" not in request.files:
        return jsonify({"error": "No file provided."}), 400

    file = request.files["file"]

    if file.mimetype != "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet":
        return jsonify({"error": "Invalid file format. Expected XLSX file."}), 400

    file_content = BytesIO(file.read())
    try:
        borders = get_xlsx_borders(file_content)
        return jsonify(borders), 200
    except Exception as e:
        return jsonify({"error": f"Error processing file: {e}"}), 500


def get_xlsx_borders(buffer: BytesIO) -> list[dict[str, int | bool]]:
    workbook = load_workbook(buffer)
    all_borders: list[dict[str, int | bool]] = []
    offset = -1

    for worksheet in workbook.worksheets:
        rows = 0

        for column in worksheet.iter_cols(max_col=1):
            for cell in column:
                # noinspection PyTypeChecker
                cell_border: Border = cell.border
                row = cell.row + offset
                has_top = cell_border.top is not None and cell_border.top.style is not None
                has_bottom = cell_border.bottom is not None and cell_border.bottom.style is not None

                all_borders.append(dict([("row", row), ("top", has_top), ("bottom", has_bottom)]))
                rows += 1

        offset += rows

    return all_borders


if __name__ == '__main__':
    app.run(host="127.0.0.1", port=getenv("PY_API_PORT"))
